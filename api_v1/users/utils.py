import io
import os
import platform
import tempfile
import zipfile

from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment
from copy import copy
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from openpyxl import load_workbook
from django.db.models import Case, When, Value, IntegerField

from django.conf import settings
from docxtpl import DocxTemplate

RU_MONTHS = {
    1: "января",
    2: "февраля",
    3: "марта",
    4: "апреля",
    5: "мая",
    6: "июня",
    7: "июля",
    8: "августа",
    9: "сентября",
    10: "октября",
    11: "ноября",
    12: "декабря",
}

EN_MONTHS = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}

FR_MONTHS = {
    1: "janvier",
    2: "février",
    3: "mars",
    4: "avril",
    5: "mai",
    6: "juin",
    7: "juillet",
    8: "août",
    9: "septembre",
    10: "octobre",
    11: "novembre",
    12: "décembre",
}


class DocumentService:
    path = str(settings.BASE_DIR) + os.sep

    def get_doc_pdf(self, context: dict, docx_temp: DocxTemplate):
        file_name = "temporary_file"
        pdf_bytes = io.BytesIO()
        docx_temp.render(context)
        docs_dir = tempfile.TemporaryDirectory(prefix=self.path)
        docx_temp.save(os.path.join(docs_dir.name, f"{file_name}.docx"))
        file_path = os.path.join(docs_dir.name, f"{file_name}.docx")
        output_dir = os.path.join(docs_dir.name)
        if platform.system().lower() == "windows":
            os.system(
                f"C:\Program^ Files\LibreOffice\program\soffice --headless --convert-to pdf --outdir {output_dir} {file_path}"
            )
        elif platform.system().lower() == "linux":
            os.system(
                f"soffice --headless --convert-to pdf --outdir {output_dir} {file_path}"
            )
        with open(
            os.path.join(docs_dir.name, f"{file_name}.pdf"), "rb"
        ) as pdf:
            pdf_bytes = io.BytesIO(pdf.read())
        return pdf_bytes

    def get_bytes_stream(self, relative_path: str):
        with open(relative_path, "rb") as file:
            file_bytes = io.BytesIO(file.read())
        return file_bytes


def write_cell(ws, row: int, col: int, value: str):
    cell = ws.cell(row=row, column=col)

    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                ws.cell(
                    row=merged_range.min_row,
                    column=merged_range.min_col
                ).value = value
                return

    cell.value = value
    
def split_text(text: str, max_len: int, max_lines: int):
    """
    Делит текст на max_lines строк, каждая не длиннее max_len,
    при этом слова не разрываются. Если слово не помещается, оно переносится на следующую строку.
    """
    if not text:
        return []

    words = text.split()
    lines = []
    current_line = ""

    for word in words:
        if current_line:
            # если текущее слово влезает с пробелом
            if len(current_line) + 1 + len(word) <= max_len:
                current_line += " " + word
            else:
                lines.append(current_line)
                current_line = word
        else:
            # первая строка
            if len(word) <= max_len:
                current_line = word
            else:
                # слово длиннее max_len, режем его
                while len(word) > max_len:
                    lines.append(word[:max_len])
                    word = word[max_len:]
                current_line = word

        if len(lines) == max_lines:
            break

    if current_line and len(lines) < max_lines:
        lines.append(current_line)

    return lines[:max_lines]


def write_merged_cell(ws, row: int, start_col: int, value: str, merge_cols: int = 1):
    if merge_cols > 1:
        end_col = start_col + merge_cols - 1
        ws.merge_cells(
            start_row=row,
            start_column=start_col,
            end_row=row,
            end_column=end_col
        )
    ws.cell(row=row, column=start_col).value = value
    
    
def insert_education_row_ru(ws, row):
    """
    Вставляет пустую строку для образования на позицию row,
    копирует стиль предыдущей строки и объединяет ячейки
    """
    # Сохраняем merge диапазоны
    merges = list(ws.merged_cells.ranges)

    # Вставляем пустую строку
    ws.insert_rows(row)

    # Сдвигаем merge диапазоны вниз
    ws.merged_cells.ranges = []
    for m in merges:
        min_col, min_row, max_col, max_row = m.bounds
        if min_row >= row:
            m.shift(0, 1)
        ws.merged_cells.add(m)

    # Копируем стиль с предыдущей строки
    for col in range(1, ws.max_column + 1):
        source = ws.cell(row=row-1, column=col)
        target = ws.cell(row=row, column=col)
        if source.has_style:
            target.font = copy(source.font)
            target.border = copy(source.border)
            target.fill = copy(source.fill)
            target.number_format = copy(source.number_format)
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)

    # Объединяем ячейки под образование
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)  # A-E
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)  # F-G
    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)  # H-I
    ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=12)  # J-L
    ws.merge_cells(start_row=row, start_column=13, end_row=row, end_column=14)  # M-N

    return row 


def insert_employment_row_ru(ws, row: int):
    """
    Вставляет строку для трудовой деятельности на позицию row.
    Копирует стиль предыдущей строки и объединяет ячейки под колонки таблицы.
    Таблица трудовой деятельности:
    Дата приема A
    Дата увольнения B-D
    Должность, наименование организации E-G
    Адрес, телефон организации H-I
    ФИО руководителя J-L
    Причина увольнения M-N
    """
    # Сохраняем merge диапазоны
    merges = list(ws.merged_cells.ranges)

    # Вставляем пустую строку
    ws.insert_rows(row)

    # Сдвигаем merge диапазоны вниз
    ws.merged_cells.ranges = []
    for m in merges:
        min_col, min_row, max_col, max_row = m.bounds
        if min_row >= row:
            m.shift(0, 1)
        ws.merged_cells.add(m)

    # Копируем стиль с предыдущей строки
    for col in range(1, ws.max_column + 1):
        source = ws.cell(row=row-1, column=col)
        target = ws.cell(row=row, column=col)
        if source.has_style:
            target.font = copy(source.font)
            target.border = copy(source.border)
            target.fill = copy(source.fill)
            target.number_format = copy(source.number_format)
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)

    # Объединяем ячейки под колонки таблицы трудовой деятельности
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)   # Дата увольнения
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)   # Должность, организация
    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)   # Адрес, телефон организации
    ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=12) # ФИО руководителя
    ws.merge_cells(start_row=row, start_column=13, end_row=row, end_column=14) # Причина увольнения
    
    
def insert_recommendation_row_ru(ws, row: int):
    """
    Вставляет строку для рекомендации на позицию row.
    Копирует стиль предыдущей строки и объединяет ячейки под всю таблицу рекомендаций (A-N)
    """
    from copy import copy

    # Сохраняем merge диапазоны
    merges = list(ws.merged_cells.ranges)

    # Вставляем пустую строку
    ws.insert_rows(row)

    # Сдвигаем merge диапазоны вниз
    ws.merged_cells.ranges = []
    for m in merges:
        min_col, min_row, max_col, max_row = m.bounds
        if min_row >= row:
            m.shift(0, 1)
        ws.merged_cells.add(m)

    # Копируем стиль с предыдущей строки
    for col in range(1, ws.max_column + 1):
        source = ws.cell(row=row-1, column=col)
        target = ws.cell(row=row, column=col)
        if source.has_style:
            target.font = copy(source.font)
            target.border = copy(source.border)
            target.fill = copy(source.fill)
            target.number_format = copy(source.number_format)
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)

    # Объединяем ячейки под всю таблицу рекомендаций A-N
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)


def write_recommendation_cell(ws, row: int, text: str, max_chars_per_line=75, line_height=11):
    """
    Записывает текст рекомендации в объединённую строку A-N,
    включает перенос текста и динамически подстраивает высоту строки.
    """
    # Объединяем ячейки A-N
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)

    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.alignment = Alignment(wrapText=True, vertical="top")

    # Оценка количества строк текста
    lines_count = sum((len(line) - 1) // max_chars_per_line + 1 for line in text.split("\n"))

    # Устанавливаем высоту строки
    ws.row_dimensions[row].height = max(line_height, lines_count * line_height)
    
    
def insert_family_row(ws, row: int):
    """
    Вставляет строку для таблицы «Состав семьи».
    Копирует стиль предыдущей строки и объединяет ячейки по шаблону:
    A-C | D-F | G-J | K-N
    """
    from copy import copy

    # Сохраняем merge диапазоны
    merges = list(ws.merged_cells.ranges)

    # Вставляем строку
    ws.insert_rows(row)

    # Корректно сдвигаем merge диапазоны
    ws.merged_cells.ranges = []
    for m in merges:
        min_col, min_row, max_col, max_row = m.bounds
        if min_row >= row:
            m.shift(0, 1)
        ws.merged_cells.add(m)

    # Копируем стиль с предыдущей строки
    for col in range(1, ws.max_column + 1):
        source = ws.cell(row=row - 1, column=col)
        target = ws.cell(row=row, column=col)
        if source.has_style:
            target.font = copy(source.font)
            target.border = copy(source.border)
            target.fill = copy(source.fill)
            target.number_format = copy(source.number_format)
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)

    # Объединения ячеек по шаблону таблицы семьи
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)    # A–C Степень родства
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)    # D–F Год рождения
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=10)   # G–J Род деятельности
    ws.merge_cells(start_row=row, start_column=11, end_row=row, end_column=14)  # K–N Место проживания


def find_row_by_text(ws, needle: str) -> int | None:
    needle = needle.lower().strip()

    for i, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            if cell.value:
                text = str(cell.value).lower().replace("\n", " ")
                if needle in text and text.startswith(needle):
                    return i
    return None


def write_answer_block(
    ws,
    question_text: str,
    value: str,
    max_lines: int,
    max_len: int = 100,
    start_col: int = 1
):
    if not value:
        return

    question_row = find_row_by_text(ws, question_text)
    if not question_row:
        return

    start_row = question_row + 1
    lines = split_text(value, max_len=max_len, max_lines=max_lines)

    # фиксируем высоту строк из шаблона
    heights = [
        ws.row_dimensions[start_row + i].height
        for i in range(max_lines)
    ]

    for i in range(max_lines):
        if i < len(lines):
            write_cell(ws, start_row + i, start_col, lines[i])
        ws.row_dimensions[start_row + i].height = heights[i]


def write_created_at(ws, candidate):
    """
    Ищет строку с кавычками в ячейке A (игнорируем пробелы) 
    и заполняет день, месяц и год.
    """
    target_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and row[0]:
            cell_val = str(row[0]).replace(" ", "")
            if cell_val.count('"') == 2:
                target_row = i
                break

    if not target_row or not candidate.created_at:
        return

    day = candidate.created_at.day
    month_num = candidate.created_at.month
    year = candidate.created_at.year

    # Месяц на русском
    month_name = RU_MONTHS[month_num]

    # Записываем
    ws.cell(row=target_row, column=1).value = f'"{day}"'
    ws.merge_cells(start_row=target_row, start_column=2, end_row=target_row, end_column=5)
    ws.cell(row=target_row, column=2).value = month_name
    ws.merge_cells(start_row=target_row, start_column=6, end_row=target_row, end_column=7)
    ws.cell(row=target_row, column=6).value = str(year)
    if candidate.signature:
        insert_signature(ws, candidate.signature.path, target_row)
    
    
def insert_signature(ws, signature_path: str, row: int, target_width: int = 100):
    """
    Вставляет подпись в указанную строку в диапазон K-N,
    масштабируя пропорционально по ширине.
    """
    # Читаем оригинальное изображение, чтобы знать его размеры
    with PILImage.open(signature_path) as pil_img:
        orig_width, orig_height = pil_img.size

    # Вычисляем коэффициент масштабирования
    scale = target_width / orig_width
    target_height = int(orig_height * scale)

    # Создаём объект openpyxl Image
    img = XLImage(signature_path)
    img.width = target_width
    img.height = target_height

    # Вставляем в ячейку K{row}
    cell_address = f'K{row-1}'
    ws.add_image(img, cell_address)
    

def insert_candidate_photo(ws, photo_path: str, start_row: int = 2, end_row: int = 9, start_col: int = 12, end_col: int = 14):
    """
    Вставляет фото кандидата в указанный диапазон (по умолчанию L2:N9),
    масштабируя пропорционально по ширине.
    """
    # Читаем оригинальное изображение
    with PILImage.open(photo_path) as pil_img:
        orig_width, orig_height = pil_img.size

    # Рассчитываем ширину диапазона в пикселях (приблизительно, 1 столбец ≈ 64px)
    target_width = (end_col - start_col + 1) * 40
    scale = target_width / orig_width
    target_height = int(orig_height * scale)

    # Создаём объект openpyxl Image
    img = XLImage(photo_path)
    img.width = target_width
    img.height = target_height

    # Вставляем в верхний левый угол диапазона
    cell_address = f"{chr(64 + start_col)}{start_row}"  # столбец в букве
    ws.add_image(img, cell_address)


def get_questionnaire_ru_xlsx(candidate, template):
    wb = load_workbook(template)
    ws = wb.active
    citizenship = candidate.citizenships.first()
    if citizenship:
        passport_value = ", ".join(filter(None, [
            citizenship.passport_series,
            citizenship.passport_number,
            f"выдан {citizenship.passport_issued_at.strftime("%d.%m.%Y")}" if citizenship.passport_issued_at else None,
            citizenship.passport_issued_by if citizenship.passport_issued_by else None
        ]))
    write_cell(ws, 2, 7, candidate.vacancy.position.name_ru) 
    write_cell(ws, 5, 3, candidate.last_name)               
    write_cell(ws, 6, 3, candidate.first_name)             
    write_cell(ws, 7, 3, candidate.middle_name)           
    write_cell(ws, 9, 3, candidate.birth_date.strftime("%d.%m.%Y") if candidate.birth_date else "")
    if citizenship:
        write_cell(ws, 10, 9, citizenship.citizenship)        
    write_cell(ws, 11, 5, candidate.birth_place)        

    if citizenship:
        passport_lines = split_text(passport_value, max_len=55, max_lines=2)
        if passport_lines:
            write_cell(ws, 12, 8, passport_lines[0])
        if len(passport_lines) > 1:
            write_cell(ws, 13, 1, passport_lines[1])

    write_cell(ws, 14, 6, candidate.phone)                  
    write_cell(ws, 15, 6, candidate.email)                    

    registration_lines = split_text(candidate.registration_address, max_len=60, max_lines=3)
    for i, line in enumerate(registration_lines):
        write_cell(ws, 16 + i, 6, line)

    residence_lines = split_text(candidate.residence_address, max_len=60, max_lines=3)
    for i, line in enumerate(residence_lines):
        write_cell(ws, 19 + i, 6, line)
    start_row = 24
    educations = candidate.educations.all().order_by("graduation_date")
    num_educations = educations.count()
    extra_rows_needed = max(0, num_educations - 4)

    for i in range(extra_rows_needed):
        insert_education_row_ru(ws, start_row + 4 + i)

    for idx, edu in enumerate(educations):
        row = start_row + idx
        ws.cell(row=row, column=1).value = edu.institution_name_and_location  
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

        ws.cell(row=row, column=6).value = edu.graduation_date
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)

        ws.cell(row=row, column=8).value = edu.get_education_form_display()
        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)

        ws.cell(row=row, column=10).value = edu.specialty
        ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=12)

        ws.cell(row=row, column=13).value = edu.diploma_information
        ws.merge_cells(start_row=row, start_column=13, end_row=row, end_column=14)

    end_of_education_row = start_row + max(num_educations, 4)
    start_employment_row = end_of_education_row + 2

    employments = candidate.employments.annotate(
        current_job=Case(
            When(end_date__isnull=True, then=Value(1)),
            default=Value(0),
            output_field=IntegerField()
        )
    ).order_by('-current_job', '-end_date', '-start_date')
    num_employments = employments.count()
    extra_rows_needed = max(0, num_employments - 7)
    for i in range(extra_rows_needed):
        insert_employment_row_ru(ws, start_employment_row + 5 + i)

    for idx, emp in enumerate(employments):
        row = start_employment_row + idx
        ws.cell(row=row, column=1).value = emp.start_date.strftime("%d.%m.%Y") if emp.start_date else ""
        ws.cell(row=row, column=2).value = emp.end_date.strftime("%d.%m.%Y") if emp.end_date else ""
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row=row, column=5).value = emp.position_and_organization
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
        ws.cell(row=row, column=8).value = emp.organization_address_and_phone
        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
        ws.cell(row=row, column=10).value = emp.manager_full_name
        ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=12)
        ws.cell(row=row, column=13).value = emp.dismissal_reason
        ws.merge_cells(start_row=row, start_column=13, end_row=row, end_column=14)
    end_of_employment_row = start_employment_row + max(num_employments, 7)

    start_foreign_languages_row = end_of_employment_row + 1 
    foreign_languages_text = candidate.foreign_languages
    lines = split_text(foreign_languages_text, max_len=75, max_lines=4)

    first_row_height = ws.row_dimensions[start_foreign_languages_row].height

    for i, line in enumerate(lines):
        write_cell(ws, start_foreign_languages_row + i, 1, line)
    for i in range(4):
        ws.row_dimensions[start_foreign_languages_row + i].height = first_row_height
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and row[0] and "Рекомендации с предыдущих мест работы" in str(row[0]):
            recommendations_start_row = i + 1 
            break
    else:
        recommendations_start_row = ws.max_row + 1

    recommendations = candidate.recommendations.all()
    num_recommendations = recommendations.count()
    extra_rows_needed = max(0, num_recommendations - 4)

    for i in range(extra_rows_needed):
        insert_recommendation_row_ru(ws, recommendations_start_row + 4 + i)

    for idx, rec in enumerate(recommendations):
        row = recommendations_start_row + idx
        write_recommendation_cell(ws, row, rec.text)
        
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and row[0] and "Состав семьи, близкие родственники" in str(row[0]):
            family_start_row = i + 2
            break
    else:
        family_start_row = ws.max_row + 1
    family_members = candidate.family_members.all()
    num_family = family_members.count()
    extra_rows_needed = max(0, num_family - 8)
    for i in range(extra_rows_needed):
        insert_family_row(ws, family_start_row + 8 + i)
    for idx, member in enumerate(family_members):
        row = family_start_row + idx
        ws.cell(row=row, column=1).value = member.relation
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.cell(row=row, column=4).value = member.birth_year
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        ws.cell(row=row, column=7).value = member.occupation
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=10)
        ws.cell(row=row, column=11).value = member.residence
        ws.merge_cells(start_row=row, start_column=11, end_row=row, end_column=14)

    write_answer_block(
        ws,
        "Являетесь ли Вы военнообязанным",
        candidate.military_service,
        max_lines=2
    )
    write_answer_block(
        ws,
        "Действует ли в отношении Вас запрет",
        candidate.disqualification,
        max_lines=2
    )
    write_answer_block(
        ws,
        "Являетесь (являлись) ли Вы руководителем",
        candidate.management_experience,
        max_lines=3
    )
    write_answer_block(
        ws,
        "Имеете ли Вы (или члены Вашей семьи) заболевания",
        candidate.health_restrictions,
        max_lines=3
    )
    driver_row = find_row_by_text(
        ws,
        "Водительское удостоверение №"
    )

    if driver_row:
        ws.cell(row=driver_row, column=8).value = candidate.driver_license_number
        ws.merge_cells(
            start_row=driver_row,
            start_column=8,
            end_row=driver_row,
            end_column=9
        )
        ws.cell(row=driver_row, column=13).value = (
            candidate.driver_license_issue_date.strftime("%d.%m.%Y")
            if candidate.driver_license_issue_date else ""
        )
        ws.merge_cells(
            start_row=driver_row,
            start_column=13,
            end_row=driver_row,
            end_column=14
        )

        categories_row = driver_row + 1
        ws.cell(row=categories_row, column=6).value = candidate.driver_license_categories
        ws.merge_cells(
            start_row=categories_row,
            start_column=6,
            end_row=categories_row,
            end_column=14
        )
    write_answer_block(
        ws,
        "Источник информации о вакансии",
        candidate.vacancy_source,
        max_lines=1
    )
    write_answer_block(
        ws,
        "Знакомые, родственники, работающие в нашей организации",
        candidate.acquaintances_in_company,
        max_lines=1
    )
    write_answer_block(
        ws,
        "Согласны ли Вы на обращение по вашему настоящему месту работы",
        (
            "Да"
            if candidate.allow_reference_check is True
            else "Нет"
            if candidate.allow_reference_check is False
            else ""
        ),
        max_lines=1
    )
    write_answer_block(
        ws,
        "Дополнительные требования к новому месту работы",
        candidate.job_requirements,
        max_lines=3
    )

    write_answer_block(
        ws,
        "Какие факторы могут стать или являются для Вас помехой в работе",
        candidate.work_obstacles,
        max_lines=3
    )
    write_answer_block(
        ws,
        "Другие сведения, которые Вы хотите сообщить о себе",
        candidate.additional_info,
        max_lines=3
    )
    write_answer_block(
        ws,
        "Ваши пожелания по заработной плате",
        candidate.salary_expectations,
        max_lines=1
    )
    write_created_at(ws, candidate)
    if candidate.photo:
        insert_candidate_photo(ws, candidate.photo.path)
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.seek(0)
    return tmp


def write_basic_info_foreign(ws, candidate):
    write_merged_cell(ws, 3, 4, candidate.last_name, merge_cols=7)   # D-J
    write_merged_cell(ws, 4, 4, candidate.first_name, merge_cols=7)
    
    if candidate.birth_date:
        write_merged_cell(
            ws, 5, 4,
            candidate.birth_date.strftime("%d.%m.%Y"),
            merge_cols=7
        )
    citizenship = candidate.citizenships.first()
    if citizenship:
        write_merged_cell(ws, 6, 4, citizenship.citizenship, merge_cols=7)
    write_merged_cell(ws, 7, 4, candidate.birth_place, merge_cols=7)
    write_merged_cell(ws, 8, 4, candidate.phone, merge_cols=7)
    write_merged_cell(ws, 9, 4, candidate.email, merge_cols=7)

    # Место проживания (10–11 строка, до 75 символов)
    lines = split_text(candidate.residence_address, max_len=75, max_lines=2)

    for i, line in enumerate(lines):
        row = 10 + i

        if i == 0:
            # первая строка: D-J
            write_merged_cell(ws, row, start_col=4, value=line, merge_cols=7)
        else:
            # вторая строка: D-N
            write_merged_cell(ws, row, start_col=4, value=line, merge_cols=11)
    # D-N
        
        
def insert_education_row_foreign(ws, row):
    """
    Вставляет пустую строку для образования на позицию row,
    копирует стиль предыдущей строки и объединяет ячейки
    """
    # Сохраняем merge диапазоны
    merges = list(ws.merged_cells.ranges)

    # Вставляем пустую строку
    ws.insert_rows(row)

    # Сдвигаем merge диапазоны вниз
    ws.merged_cells.ranges = []
    for m in merges:
        min_col, min_row, max_col, max_row = m.bounds
        if min_row >= row:
            m.shift(0, 1)
        ws.merged_cells.add(m)

    # Копируем стиль с предыдущей строки
    for col in range(1, ws.max_column + 1):
        source = ws.cell(row=row-1, column=col)
        target = ws.cell(row=row, column=col)
        if source.has_style:
            target.font = copy(source.font)
            target.border = copy(source.border)
            target.fill = copy(source.fill)
            target.number_format = copy(source.number_format)
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)

    # Объединяем ячейки под образование
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)  # A-E
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)  # F-G
    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=12)  # H-L
    ws.merge_cells(start_row=row, start_column=13, end_row=row, end_column=14)  # M-N

    return row 
        

def write_education(ws, candidate, search_string):
    header_row = find_row_by_text(ws, search_string)
    if not header_row:
        return

    start_row = header_row + 2
    educations = list(candidate.educations.all())

    base_rows = 4
    if len(educations) > base_rows:
        for i in range(len(educations) - base_rows):
            insert_education_row_foreign(ws, start_row + base_rows + i)

    for i, edu in enumerate(educations):
        row = start_row + i
        write_cell(ws, row, 1, edu.institution_name_and_location)  # A-E
        write_cell(ws, row, 6, str(edu.graduation_date))           # F-G
        write_cell(ws, row, 8, edu.specialty)                      # H-L
        write_cell(ws, row, 13, edu.diploma_information)           # M-N
        
        
def insert_employment_row_foreign(ws, row: int):
    merges = list(ws.merged_cells.ranges)

    # Вставляем пустую строку
    ws.insert_rows(row)

    # Сдвигаем merge диапазоны вниз
    ws.merged_cells.ranges = []
    for m in merges:
        min_col, min_row, max_col, max_row = m.bounds
        if min_row >= row:
            m.shift(0, 1)
        ws.merged_cells.add(m)

    # Копируем стиль с предыдущей строки
    for col in range(1, ws.max_column + 1):
        source = ws.cell(row=row-1, column=col)
        target = ws.cell(row=row, column=col)
        if source.has_style:
            target.font = copy(source.font)
            target.border = copy(source.border)
            target.fill = copy(source.fill)
            target.number_format = copy(source.number_format)
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)

    # Объединяем ячейки под колонки таблицы трудовой деятельности
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)   
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=12)   
    ws.merge_cells(start_row=row, start_column=13, end_row=row, end_column=14) 


def write_employment(ws, candidate, search_string):
    header_row = find_row_by_text(ws, search_string)
    if not header_row:
        return

    start_row = header_row + 2
    employments = candidate.employments.annotate(
        current_job=Case(
            When(end_date__isnull=True, then=Value(1)),
            default=Value(0),
            output_field=IntegerField()
        )
    ).order_by('-current_job', '-end_date', '-start_date')

    base_rows = 6
    if len(employments) > base_rows:
        for i in range(len(employments) - base_rows):
            insert_employment_row_foreign(ws, start_row + base_rows + i)

    for i, emp in enumerate(employments):
        row = start_row + i

        period = ""
        if emp.start_date:
            period += f"{emp.start_date.strftime("%m.%Y")} - "
        if emp.end_date:
            period += emp.end_date.strftime('%m.%Y')

        write_cell(ws, row, 1, period)                              # A-B
        write_cell(ws, row, 5, emp.position_and_organization)       # C-F
        write_cell(ws, row, 8, emp.organization_address_and_phone)  # G-L
        write_cell(ws, row, 13, emp.dismissal_reason)               # M-N
        
        
def insert_recommendation_row_foreign(ws, row: int):
    """
    Вставляет строку для рекомендации на позицию row.
    Копирует стиль предыдущей строки и объединяет ячейки под всю таблицу рекомендаций (A-N)
    """
    from copy import copy

    # Сохраняем merge диапазоны
    merges = list(ws.merged_cells.ranges)

    # Вставляем пустую строку
    ws.insert_rows(row)

    # Сдвигаем merge диапазоны вниз
    ws.merged_cells.ranges = []
    for m in merges:
        min_col, min_row, max_col, max_row = m.bounds
        if min_row >= row:
            m.shift(0, 1)
        ws.merged_cells.add(m)

    # Копируем стиль с предыдущей строки
    for col in range(1, ws.max_column + 1):
        source = ws.cell(row=row-1, column=col)
        target = ws.cell(row=row, column=col)
        if source.has_style:
            target.font = copy(source.font)
            target.border = copy(source.border)
            target.fill = copy(source.fill)
            target.number_format = copy(source.number_format)
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)

    # Объединяем ячейки под всю таблицу рекомендаций A-N
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
    ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=11)
    ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=14)


def write_references(ws, candidate, search_string):
    header_row = find_row_by_text(
        ws,
        search_string
    )
    if not header_row:
        return

    start_row = header_row + 2
    refs = list(candidate.recommendations.all())

    base_rows = 3
    if len(refs) > base_rows:
        for i in range(len(refs) - base_rows):
            insert_recommendation_row_foreign(ws, start_row + base_rows + i)

    for i, ref in enumerate(refs):
        row = start_row + i
        write_cell(ws, row, 1, ref.company)             
        write_cell(ws, row, 5, ref.name)       
        write_cell(ws, row, 9, ref.position)  
        write_cell(ws, row, 12, ref.contact) 
        
        
def write_family(ws, candidate, search_string):
    header_row = find_row_by_text(ws, search_string)
    if not header_row:
        return

    start_row = header_row + 2
    family = list(candidate.family_members.all())

    base_rows = 6
    if len(family) > base_rows:
        for i in range(len(family) - base_rows):
            insert_family_row(ws, start_row + base_rows + i)

    for i, member in enumerate(family):
        row = start_row + i

        write_cell(ws, row, 1, member.relation)
        write_cell(ws, row, 4, member.birth_date.strftime("%d.%m.%Y") if member.birth_date else "")
        write_cell(ws, row, 7, member.occupation)
        write_cell(ws, row, 11, member.residence)
         

def write_questions_en(ws, candidate):
    write_answer_block(
        ws,
        "Do you have any occupational health problems or a disability?",
        candidate.health_restrictions,
        max_lines=2
    )
    write_answer_block(ws, "Where did you find out about us?", candidate.vacancy_source, 1)
    write_answer_block(
        ws,
        "Do you currently have any friends, relatives working for us?",
        candidate.acquaintances_in_company,
        2
    )
    write_answer_block(ws, "Do you have any additional job requirements?", candidate.job_requirements, 2)
    write_answer_block(
        ws,
        "Please indicate which factors can be disturbance in your work:",
        candidate.work_obstacles,
        2
    )
    write_answer_block(
        ws,
        "Other personal details you are willing to give:",
        candidate.additional_info,
        3
    )
    write_answer_block(ws, "Your salary expectations:", candidate.salary_expectations, 1)


def get_questionnaire_en_xlsx(candidate, template):
    wb = load_workbook(template)
    ws = wb.active
    write_basic_info_foreign(ws, candidate)
    write_education(ws, candidate, "education")
    write_employment(ws, candidate, "employment history")
    write_references(ws, candidate, "references (may be provided by your manager or business partners)")
    write_family(ws, candidate, "family")
    write_questions_en(ws, candidate)
    write_created_at(ws, candidate)

    if candidate.photo:
        insert_candidate_photo(ws, candidate.photo.path, start_row=3)
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.seek(0)
    return tmp


def write_questions_fr(ws, candidate):
    write_answer_block(
        ws,
        "Avez-vous des maladies professionnelles ou un handicap?",
        candidate.health_restrictions,
        max_lines=2
    )
    write_answer_block(ws, "Comment vous avez appris de notre poste?", candidate.vacancy_source, 1)
    write_answer_block(
        ws,
        "Avez-vous des amis ou membres de famille qui travaillent dans notre compagnie?",
        candidate.acquaintances_in_company,
        2
    )
    write_answer_block(ws, "Avez-vous des demandes supplémentaires?", candidate.job_requirements, 2)
    write_answer_block(
        ws,
        "Notez, s'il vous plait, les facteurs qui peuvent être des soucis pour le travail:",
        candidate.work_obstacles,
        2
    )
    write_answer_block(
        ws,
        "Autre information de vous-même que vous voulez partager:",
        candidate.additional_info,
        3
    )
    write_answer_block(ws, "Niveau de salaire expecté:", candidate.salary_expectations, 1)


def get_questionnaire_fr_xlsx(candidate, template):
    wb = load_workbook(template)
    ws = wb.active
    write_basic_info_foreign(ws, candidate)
    write_education(ws, candidate, "formation")
    write_employment(ws, candidate, "expérience professionelle")
    write_references(ws, candidate, "références (chefs ou business partenaires)")
    write_family(ws, candidate, "famille")
    write_questions_fr(ws, candidate)
    write_created_at(ws, candidate)

    if candidate.photo:
        insert_candidate_photo(ws, candidate.photo.path, start_row=3)
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.seek(0)
    return tmp